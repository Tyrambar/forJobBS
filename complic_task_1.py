import re
import string as st
from random import choice
from csv import writer

import requests
from bs4 import BeautifulSoup as Bs
from openpyxl import Workbook

from for_task_1 import *


PARAMS_COMPANY[0] = "№"
companies = []
while True:
    inp_url = input().lower()
    if re.match(r'({})\d+'.format(main_url), inp_url):
        appended_url = re.sub('{}'.format(main_url), '', inp_url)
        companies.append(re.sub(' ', '', appended_url))
    if inp_url.lower() == 'stop':
        break

HEAD_CELLS = ["{}{}".format(st.ascii_uppercase[numb], 1) 
    for numb in range(len(PARAMS_COMPANY))]
wb = Workbook()
ws = wb.active

for numb, cell in enumerate(HEAD_CELLS):
    ws[cell] = PARAMS_COMPANY[numb]

	
PATH = "for_task_1_complic.csv"
with open(PATH, "w", newline='') as for_task_1_complic:
    write = writer(for_task_1_complic, delimiter=",")
    write.writerow(PARAMS_COMPANY)
    for numb, company in enumerate(companies):
        for i in range(20):
            USERAGENT = {'User-agent': choice(USERAGENTS)}
            try:
                req = requests.get('{}{}'.format(main_url, company),
                                   headers=USERAGENT)
            except:
                pass
            else:
                break
        bs = Bs(req.text, "html.parser")
        ans = get_params(bs)
        main_cells = ["{}{}".format(st.ascii_uppercase[num+1], numb+2)
            for num in range(len(PARAMS_COMPANY[1:]))]
        ws["A" + str(numb+2)] = numb+1
        row_csv = [numb+1]
        for num, cell in enumerate(main_cells):
            ws[cell] = ans[PARAMS_COMPANY[num+1]]
            row_csv.append(ans[PARAMS_COMPANY[num+1]])
        write.writerow(row_csv)

		
wb.save("for_task_1_complic.xlsx")
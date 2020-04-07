import re
import string as st
from random import choice
from csv import writer

import requests
from bs4 import BeautifulSoup as Bs
from openpyxl import Workbook

from for_task_1 import *


PARAMS_COMPANY[0] = "Характеристики компании"
while True:
    inp_url = input().lower()
    if re.match(r'\d+', inp_url):
        break

		
HEAD_CELLS = ["{}{}".format('A', numb+1) 
	for numb in range(len(PARAMS_COMPANY))]
wb = Workbook()
ws = wb.active
for numb, cell in enumerate(HEAD_CELLS):
    ws[cell] = PARAMS_COMPANY[numb]
url = '{}{}'.format(main_url, inp_url)

for i in range(20):
    USERAGENT = {'User-agent': choice(USERAGENTS)}
    try:
        req = requests.get(url, headers=USERAGENT)
    except:
        pass
    else:
        break

		
bs = Bs(req.text, "html.parser")
ans = get_params(bs)

for n, param in enumerate(PARAMS_COMPANY[1:]):
    main_cells = ["{}{}".format('B', numb+2) 
		for numb in range(len(PARAMS_COMPANY[1:]))]
    for num, cell in enumerate(main_cells):
        ws[cell] = ans[PARAMS_COMPANY[num+1]]

		
PATH = "for_task_1_simple.csv"

with open(PATH, "w", newline='') as for_task_1_simple:
    write = writer(for_task_1_simple, delimiter=",")
    write.writerow([PARAMS_COMPANY[0],''])
    for line in range(len(PARAMS_COMPANY[:-1])):
        write.writerow([PARAMS_COMPANY[line+1],
					    ans[PARAMS_COMPANY[line+1]]])

wb.save("for_task_1_simple.xlsx")



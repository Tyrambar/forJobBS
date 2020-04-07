import re
import string as st
from random import choice
from csv import writer

import requests
from bs4 import BeautifulSoup as Bs
from openpyxl import Workbook

from for_task_1 import USERAGENTS


URL = "https://habr.com/ru/top/yearly/"

PARAMS_ARTICLES = ['№', 'заголовок поста',
                   'короткое описание поста', 
                   'дата публикации', 'имя автора поста']
HEAD_CELLS = ["{}{}".format(st.ascii_uppercase[numb], 1)
    for numb in range(len(PARAMS_ARTICLES))]
wb = Workbook()
ws = wb.active

for numb, cell in enumerate(HEAD_CELLS):
    ws[cell] = PARAMS_ARTICLES[numb]
	
	
def get_request(n_pages = 0):
    additional_url = ""
    if n_pages:
        additional_url = "page" + str(n_pages)
    for i in range(10):
        USERAGENT = {'User-agent': choice(USERAGENTS)}
        try:
            req = requests.get("{}{}".format(url, additional_url),
                               headers=USERAGENT)
        except:
            pass
        else:
            return req
			
			
count_art = int(input())
previous = ''
ans = {}
all_articles = []
counter = 2

while len(all_articles) != count_art:
    req = get_request(counter if all_articles else 0)
    bs = Bs(req.text, "html.parser")
    from_page_articles = bs.find_all("li", 
                                     class_="content-list__item_post")
    for art in from_page_articles:
        str_art = str(art)
        if 'id="post' in str_art:
            all_articles.append(art)
            if len(all_articles) == count_art:
                break
				

PATH = "for_task_2.csv"
with open(PATH, "w", newline='') as for_task_2:
    write = writer(for_task_2, delimiter="|")
    write.writerow(PARAMS_ARTICLES)
	
    for numb, art in enumerate(all_articles):
        ans[PARAMS_ARTICLES[0]] = numb+1
        ans[PARAMS_ARTICLES[1]] = art.find("a", class_="post__title_link")
                                           .get_text()
        ans[PARAMS_ARTICLES[2]] = art.find("div", class_="post__text-html")
                                           .get_text()
        ans[PARAMS_ARTICLES[3]] = art.find("span", 
                                           class_="post__time").get_text()
        ans[PARAMS_ARTICLES[4]] = art.find("span",
                                           class_ = "user-info__nickname")
                                           .get_text()
        main_cells = ["{}{}".format(st.ascii_uppercase[num], numb+2)
            for num in range(len(PARAMS_ARTICLES))]
			
        row_csv = []
        for num, cell in enumerate(main_cells):
            ws[cell] = ans[PARAMS_ARTICLES[num]]
            row_csv.append(ans[PARAMS_ARTICLES[num]])
        write.writerow(row_csv)
		
wb.save("for_task_2.xlsx")